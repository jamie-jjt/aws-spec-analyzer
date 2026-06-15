"""
exporter.py - Generates CSV and Excel exports of the BOM.
"""

import csv
import io
import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


BOM_COLUMNS = [
    "Category",
    "AWS Service",
    "Service Type / SKU",
    "Quantity",
    "Unit",
    "Pricing Model",
    "Monthly Cost (USD)",
    "Annual Cost (USD)",
    "Notes / Source Spec",
    "AWS Calculator URL",
]


def build_bom_rows(mappings: list, project_name: str = "AWS BOM") -> list:
    """Convert mapping dicts (as edited by user) into flat BOM rows."""
    rows = []
    for m in mappings:
        monthly = float(m.get("selected_monthly_usd", m.get("monthly_estimate_usd", 0)))
        qty = int(m.get("selected_quantity", m.get("quantity", 1)))
        total_monthly = monthly  # already includes quantity from mapping
        rows.append({
            "Category": m.get("category", ""),
            "AWS Service": m.get("service_name", ""),
            "Service Type / SKU": m.get("selected_type", m.get("recommended_type", "")),
            "Quantity": qty,
            "Unit": m.get("unit", ""),
            "Pricing Model": m.get("selected_pricing", "On-Demand"),
            "Monthly Cost (USD)": round(total_monthly, 2),
            "Annual Cost (USD)": round(total_monthly * 12, 2),
            "Notes / Source Spec": m.get("notes", ""),
            "AWS Calculator URL": m.get("aws_calculator_url", ""),
        })
    return rows


def export_csv(mappings: list, project_name: str = "AWS BOM", region: str = "us-east-1") -> bytes:
    """Return CSV bytes for the BOM."""
    rows = build_bom_rows(mappings, project_name)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=BOM_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

    # Summary row
    total_monthly = sum(r["Monthly Cost (USD)"] for r in rows)
    total_annual = sum(r["Annual Cost (USD)"] for r in rows)
    writer.writerow({
        "Category": "TOTAL",
        "AWS Service": "",
        "Service Type / SKU": "",
        "Quantity": "",
        "Unit": "",
        "Pricing Model": "",
        "Monthly Cost (USD)": round(total_monthly, 2),
        "Annual Cost (USD)": round(total_annual, 2),
        "Notes / Source Spec": f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | Region: {region}",
        "AWS Calculator URL": "https://calculator.aws/pricing/2/home",
    })

    return output.getvalue().encode("utf-8")


def export_excel(mappings: list, project_name: str = "AWS BOM", region: str = "us-east-1") -> bytes:
    """Return Excel (.xlsx) bytes for the BOM — styled like an official pricing sheet."""
    rows = build_bom_rows(mappings, project_name)
    wb = openpyxl.Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws_bom = wb.active
    ws_bom.title = "AWS BOM"

    # Colors
    AWS_ORANGE = "FF9900"
    AWS_DARK = "232F3E"
    HEADER_BG = "232F3E"
    ALT_ROW = "F5F5F5"
    WHITE = "FFFFFF"
    TOTAL_BG = "FFF3CD"
    BORDER_COLOR = "CCCCCC"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # ── Title block ──────────────────────────────────────────────────────────
    ws_bom.merge_cells("A1:J1")
    title_cell = ws_bom["A1"]
    title_cell.value = f"AWS Bill of Materials — {project_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=AWS_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_bom.row_dimensions[1].height = 30

    ws_bom.merge_cells("A2:J2")
    sub_cell = ws_bom["A2"]
    sub_cell.value = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y %H:%M')}  |  All prices in USD  |  Region: {region}"
    sub_cell.font = Font(name="Calibri", italic=True, size=10, color="666666")
    sub_cell.fill = PatternFill("solid", fgColor="F8F8F8")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_bom.row_dimensions[2].height = 18

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = 4
    ws_bom.row_dimensions[header_row].height = 22
    for col_idx, col_name in enumerate(BOM_COLUMNS, start=1):
        cell = ws_bom.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=header_row + 1):
        bg = ALT_ROW if row_idx % 2 == 0 else WHITE
        ws_bom.row_dimensions[row_idx].height = 18
        for col_idx, col_name in enumerate(BOM_COLUMNS, start=1):
            val = row.get(col_name, "")
            cell = ws_bom.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 9))

            # Format numbers
            if col_name in ("Monthly Cost (USD)", "Annual Cost (USD)"):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Quantity":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Totals row ────────────────────────────────────────────────────────────
    total_row = header_row + len(rows) + 1
    ws_bom.row_dimensions[total_row].height = 22
    total_monthly = sum(r["Monthly Cost (USD)"] for r in rows)
    total_annual = sum(r["Annual Cost (USD)"] for r in rows)

    totals = {
        "Category": "TOTAL",
        "Monthly Cost (USD)": round(total_monthly, 2),
        "Annual Cost (USD)": round(total_annual, 2),
    }

    for col_idx, col_name in enumerate(BOM_COLUMNS, start=1):
        val = totals.get(col_name, "")
        cell = ws_bom.cell(row=total_row, column=col_idx, value=val)
        cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
        cell.border = thin_border
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.alignment = Alignment(vertical="center")
        if col_name in ("Monthly Cost (USD)", "Annual Cost (USD)"):
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center", bold=True)

    # ── Summary cost box ──────────────────────────────────────────────────────
    summary_row = total_row + 2
    ws_bom.merge_cells(f"H{summary_row}:J{summary_row}")
    cell = ws_bom.cell(row=summary_row, column=7, value="Estimated Monthly Total:")
    cell.font = Font(name="Calibri", bold=True, size=12, color=AWS_DARK)
    cell.alignment = Alignment(horizontal="right")

    cell2 = ws_bom.cell(row=summary_row, column=8, value=round(total_monthly, 2))
    cell2.font = Font(name="Calibri", bold=True, size=14, color=AWS_DARK)
    cell2.number_format = '"$"#,##0.00'
    cell2.fill = PatternFill("solid", fgColor=AWS_ORANGE)

    ws_bom.merge_cells(f"H{summary_row+1}:J{summary_row+1}")
    cell3 = ws_bom.cell(row=summary_row + 1, column=7, value="Estimated Annual Total:")
    cell3.font = Font(name="Calibri", bold=True, size=12, color=AWS_DARK)
    cell3.alignment = Alignment(horizontal="right")

    cell4 = ws_bom.cell(row=summary_row + 1, column=8, value=round(total_annual, 2))
    cell4.font = Font(name="Calibri", bold=True, size=14, color=AWS_DARK)
    cell4.number_format = '"$"#,##0.00'
    cell4.fill = PatternFill("solid", fgColor="FFD580")

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [16, 28, 30, 10, 12, 20, 20, 20, 40, 45]
    for i, w in enumerate(col_widths, start=1):
        ws_bom.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws_bom.freeze_panes = f"A{header_row + 1}"

    # ── Details Sheet ─────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet(title="Service Details")

    detail_headers = ["Category", "AWS Service", "Recommended Type", "Alternatives",
                      "Confidence", "Missing Information", "Notes"]
    for col_idx, h in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_detail.row_dimensions[1].height = 22

    for row_idx, m in enumerate(mappings, start=2):
        alts = "; ".join(
            f"{a.get('label', '')} ({a.get('type', '')})"
            for a in m.get("alternatives", [])
        )
        missing = "; ".join(m.get("missing_info", []))
        detail_row = [
            m.get("category", ""),
            m.get("service_name", ""),
            m.get("selected_type", m.get("recommended_type", "")),
            alts,
            m.get("confidence", ""),
            missing,
            m.get("notes", ""),
        ]
        bg = ALT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(detail_row, start=1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_detail.row_dimensions[row_idx].height = 40

    detail_widths = [16, 28, 30, 60, 12, 50, 40]
    for i, w in enumerate(detail_widths, start=1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w
    ws_detail.freeze_panes = "A2"

    # ── Disclaimer Sheet ──────────────────────────────────────────────────────
    ws_disc = wb.create_sheet(title="Disclaimer")
    disclaimers = [
        ("AWS BOM Pricing Disclaimer", True, 14),
        ("", False, 11),
        ("1. All prices shown are estimates based on public AWS pricing as of the document generation date.", False, 11),
        ("2. Prices are in USD and reflect on-demand rates for the us-east-1 (N. Virginia) region unless otherwise noted.", False, 11),
        ("3. Actual costs may vary based on region, usage patterns, volume discounts, enterprise agreements, and AWS price changes.", False, 11),
        ("4. This BOM does not account for data transfer costs, support plan costs, or taxes.", False, 11),
        ("5. Reserved Instance and Savings Plans discounts are approximate. Consult AWS for exact commitments.", False, 11),
        ("6. AWS Graviton pricing reflects approximately 20% savings vs. equivalent x86 instances.", False, 11),
        ("7. Always verify pricing with the official AWS Pricing Calculator: https://calculator.aws/pricing/2/home", False, 11),
        ("", False, 11),
        ("For official pricing, visit: https://aws.amazon.com/pricing/", False, 11),
    ]
    for row_idx, (text, bold, size) in enumerate(disclaimers, start=1):
        cell = ws_disc.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True)
    ws_disc.column_dimensions["A"].width = 120

    # ── Save ──────────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
